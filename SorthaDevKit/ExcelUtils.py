import pandas as pd
import openpyxl
from typing import List, Dict, Any
from .StateBase import QuestionAnswer, ExcelOutputType, AzureMigrateServer, AzureMigrateReport, DependencyAnalysis, DependencyConnection, NetworkSegment
import os
from datetime import datetime
import re

class ExcelProcessor:
    """Utility class for processing Excel files with questions and generating output Excel files."""
    
    @staticmethod
    def read_questions_from_excel(file_path: str, question_column: str = 'Questions', sheet_name: str = None) -> List[Dict[str, str]]:
        """
        Read questions from an Excel file with Questions, Category, and Priority columns.
        
        Args:
            file_path: Path to the Excel file
            question_column: Name of the column containing questions
            sheet_name: Name of the sheet to read (None for first sheet)
            
        Returns:
            List of dictionaries with question, category, and priority
        """
        try:
            # Determine which engine to use based on file extension
            # xlrd 2.0+ only supports .xls files, openpyxl handles .xlsx
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.xlsx':
                # For .xlsx files, use openpyxl engine
                engines = ['openpyxl']
            elif file_ext == '.xls':
                # For .xls files, use xlrd engine
                engines = ['xlrd']
            else:
                # For other extensions, try both
                engines = ['openpyxl', 'xlrd']
            
            df = None
            last_error = None
            
            for engine in engines:
                try:
                    if sheet_name:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
                    else:
                        df = pd.read_excel(file_path, engine=engine)
                    break  # Success, break out of loop
                except Exception as e:
                    last_error = e
                    continue
            
            if df is None:
                # If specific engines failed, try without specifying engine (let pandas decide)
                try:
                    if sheet_name:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                    else:
                        df = pd.read_excel(file_path)
                except Exception as e:
                    raise Exception(f"Failed to read Excel file {file_path}. File extension: {file_ext}. Last error: {str(last_error)}")
            
            # Handle different possible column names for Questions
            possible_question_columns = [question_column, 'Questions', 'Question', 'questions', 'QUESTION']
            question_col = None
            
            for col in possible_question_columns:
                if col in df.columns:
                    question_col = col
                    break
            
            if question_col is None:
                # If no standard column found, use the first column
                question_col = df.columns[0]
            
            # Handle different possible column names for Category
            possible_category_columns = ['Category', 'category', 'CATEGORY', 'Categories']
            category_col = None
            
            for col in possible_category_columns:
                if col in df.columns:
                    category_col = col
                    break
            
            # Handle different possible column names for Priority
            possible_priority_columns = ['Priority', 'priority', 'PRIORITY', 'Priorities']
            priority_col = None
            
            for col in possible_priority_columns:
                if col in df.columns:
                    priority_col = col
                    break
            
            # Extract questions with their metadata
            questions_data = []
            for index, row in df.iterrows():
                question = str(row[question_col]).strip() if pd.notna(row[question_col]) else ""
                if question:  # Only add non-empty questions
                    category = str(row[category_col]).strip() if category_col and pd.notna(row[category_col]) else "General"
                    priority = str(row[priority_col]).strip() if priority_col and pd.notna(row[priority_col]) else "Medium"
                    
                    questions_data.append({
                        'question': question,
                        'category': category,
                        'priority': priority
                    })
            
            return questions_data
            
        except Exception as e:
            raise Exception(f"Error reading Excel file {file_path}: {str(e)}")
    
    @staticmethod
    def read_azure_migrate_report(file_path: str) -> AzureMigrateReport:
        """
        Read and parse Azure Migrate report from Excel file.
        
        Args:
            file_path: Path to the Azure Migrate Excel report
            
        Returns:
            AzureMigrateReport object with parsed data
        """
        try:
            # Determine which engine to use based on file extension
            # xlrd 2.0+ only supports .xls files, openpyxl handles .xlsx
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.xlsx':
                # For .xlsx files, use openpyxl engine
                engines = ['openpyxl']
            elif file_ext == '.xls':
                # For .xls files, use xlrd engine
                engines = ['xlrd']
            else:
                # For other extensions, try both
                engines = ['openpyxl', 'xlrd']
            
            excel_file = None
            last_error = None
            
            for engine in engines:
                try:
                    excel_file = pd.ExcelFile(file_path, engine=engine)
                    break
                except Exception as e:
                    last_error = e
                    continue
            
            if excel_file is None:
                try:
                    excel_file = pd.ExcelFile(file_path)
                except Exception as e:
                    raise Exception(f"Failed to read Excel file {file_path}. File extension: {file_ext}. Last error: {str(last_error)}")
            
            servers = []
            summary = {}
            metadata = {"source_file": file_path, "sheets_processed": []}
            
            # Process each sheet
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, engine=excel_file.engine)
                    metadata["sheets_processed"].append(sheet_name)
                    
                    # Try to identify server data sheets
                    if ExcelProcessor._is_server_data_sheet(df, sheet_name):
                        sheet_servers = ExcelProcessor._parse_server_sheet(df, sheet_name)
                        servers.extend(sheet_servers)
                    
                    # Try to identify summary sheets
                    elif ExcelProcessor._is_summary_sheet(df, sheet_name):
                        sheet_summary = ExcelProcessor._parse_summary_sheet(df, sheet_name)
                        summary.update(sheet_summary)
                        
                except Exception as e:
                    metadata[f"error_{sheet_name}"] = str(e)
                    continue
            
            return AzureMigrateReport(
                servers=servers,
                summary=summary,
                metadata=metadata
            )
            
        except Exception as e:
            raise Exception(f"Error reading Azure Migrate report {file_path}: {str(e)}")
    
    @staticmethod
    def _is_server_data_sheet(df: pd.DataFrame, sheet_name: str) -> bool:
        """Check if the sheet contains server/machine data."""
        sheet_name_lower = sheet_name.lower()
        
        # Specific Azure Migrate sheet names
        azure_migrate_machine_sheets = [
            'all assessed machines', 'assessed machines', 'machines', 
            'servers', 'vm assessment', 'server assessment'
        ]
        
        # Check for exact Azure Migrate sheet names
        if any(sheet_name_lower == indicator or indicator in sheet_name_lower 
               for indicator in azure_migrate_machine_sheets):
            return True
        
        # Check for general server indicators
        server_indicators = ['server', 'machine', 'vm', 'computer', 'host', 'node']
        if any(indicator in sheet_name_lower for indicator in server_indicators):
            return True
        
        # Check column headers for machine data indicators
        if len(df.columns) > 0:
            columns_str = ' '.join(df.columns.astype(str)).lower()
            column_indicators = [
                'server name', 'machine name', 'computer name', 'hostname', 
                'operating system', 'cpu', 'memory', 'disk', 'recommendation',
                'azure vm size', 'azure readiness', 'monthly cost estimate'
            ]
            if any(indicator in columns_str for indicator in column_indicators):
                return True
        
        return False
    
    @staticmethod
    def _is_summary_sheet(df: pd.DataFrame, sheet_name: str) -> bool:
        """Check if the sheet contains summary data."""
        sheet_name_lower = sheet_name.lower()
        
        # Specific Azure Migrate summary sheet names
        azure_migrate_summary_sheets = [
            'assessment summary', 'summary', 'overview', 'assessment properties'
        ]
        
        # Check for exact Azure Migrate sheet names
        if any(sheet_name_lower == indicator or indicator in sheet_name_lower 
               for indicator in azure_migrate_summary_sheets):
            return True
        
        # Check for general summary indicators
        summary_indicators = ['total', 'cost', 'recommendation', 'properties']
        return any(indicator in sheet_name_lower for indicator in summary_indicators)
    
    @staticmethod
    def _parse_server_sheet(df: pd.DataFrame, sheet_name: str) -> List[AzureMigrateServer]:
        """Parse server data from a sheet."""
        servers = []
        
        # Map common column variations to standard names (Azure Migrate specific)
        column_mapping = {
            'server_name': [
                'machine', 'server name', 'machine name', 'computer name', 'hostname', 'name', 'servername',
                'display name', 'vm name'
            ],
            'server_type': [
                'server type', 'machine type', 'type', 'servertype', 'operating system type',
                'platform', 'vm type', 'vm host'
            ],
            'operating_system': [
                'operating system', 'os', 'operatingsystem', 'platform', 'os name',
                'operating system name', 'os version'
            ],
            'cpu_cores': [
                'cores', 'cpu cores', 'processor cores', 'cpucores', 'vcpus', 'logical processors',
                'number of cores', 'core count', 'processors', 'processor'
            ],
            'memory_gb': [
                'memory(mb)', 'memory (mb)', 'memory(gb)', 'memory (gb)', 'memory', 'ram', 'ram (gb)', 'memory gb',
                'total memory', 'physical memory', 'memory size'
            ],
            'disk_size_gb': [
                'storage(gb)', 'storage (gb)', 'disk size (gb)', 'disk', 'storage', 'disk space', 'disk gb',
                'total disk size', 'storage size', 'disk capacity'
            ],
            'network_adapters': [
                'network adapters', 'nics', 'network cards', 'network interfaces',
                'ethernet adapters', 'network adapter count'
            ],
            'recommendation': [
                'recommended size', 'recommendation', 'azure recommendation', 'suggested sku', 'azure vm size',
                'recommended size', 'vm size recommendation', 'azure vm recommendation'
            ],
            'readiness': [
                'azure vm readiness', 'azure readiness', 'readiness', 'migration readiness', 'ready',
                'ready for azure', 'assessment status'
            ],
            'estimated_cost': [
                'compute monthly cost estimate usd', 'estimated cost', 'cost', 'monthly cost', 'cost estimate', 'monthly cost estimate',
                'azure cost', 'monthly cost (usd)', 'estimated monthly cost'
            ],
            'confidence': [
                'confidence rating (% of utilization data collected)', 'confidence', 'confidence rating', 'assessment confidence', 'rating confidence'
            ],
            'azure_vm_size': [
                'recommended size', 'azure vm size', 'vm size', 'recommended vm size', 'target vm size'
            ],
            'storage_type': [
                'storage type', 'disk type', 'recommended storage', 'azure storage type'
            ],
            'boot_type': [
                'boot type', 'boot', 'startup type'
            ],
            'cpu_usage': [
                'cpu usage(%)', 'cpu usage', 'processor usage', 'cpu utilization'
            ],
            'memory_usage': [
                'memory usage(%)', 'memory usage', 'ram usage', 'memory utilization'
            ]
        }
        
        # Find actual column names
        df_columns_lower = [col.lower().strip() for col in df.columns]
        mapped_columns = {}
        
        for standard_name, variations in column_mapping.items():
            for variation in variations:
                if variation in df_columns_lower:
                    mapped_columns[standard_name] = df.columns[df_columns_lower.index(variation)]
                    break
        
        # Process each row
        for _, row in df.iterrows():
            try:
                server = AzureMigrateServer()
                
                # Extract server information using mapped columns
                if 'server_name' in mapped_columns:
                    server.server_name = str(row[mapped_columns['server_name']]) if pd.notna(row[mapped_columns['server_name']]) else ""
                
                if 'server_type' in mapped_columns:
                    server.server_type = str(row[mapped_columns['server_type']]) if pd.notna(row[mapped_columns['server_type']]) else ""
                
                if 'operating_system' in mapped_columns:
                    server.operating_system = str(row[mapped_columns['operating_system']]) if pd.notna(row[mapped_columns['operating_system']]) else ""
                
                if 'cpu_cores' in mapped_columns:
                    try:
                        server.cpu_cores = int(float(str(row[mapped_columns['cpu_cores']]).replace(',', ''))) if pd.notna(row[mapped_columns['cpu_cores']]) else 0
                    except:
                        server.cpu_cores = 0
                
                if 'memory_gb' in mapped_columns:
                    try:
                        memory_val = str(row[mapped_columns['memory_gb']]).replace(',', '').replace('GB', '').replace('MB', '').strip()
                        server.memory_gb = float(memory_val) if memory_val else 0.0
                        # Convert MB to GB if needed
                        if 'mb' in mapped_columns['memory_gb'].lower():
                            server.memory_gb = server.memory_gb / 1024
                    except:
                        server.memory_gb = 0.0
                
                if 'disk_size_gb' in mapped_columns:
                    try:
                        disk_val = str(row[mapped_columns['disk_size_gb']]).replace(',', '').replace('GB', '').replace('TB', '').strip()
                        server.disk_size_gb = float(disk_val) if disk_val else 0.0
                        # Convert TB to GB if needed
                        if 'tb' in mapped_columns['disk_size_gb'].lower():
                            server.disk_size_gb = server.disk_size_gb * 1024
                    except:
                        server.disk_size_gb = 0.0
                
                if 'network_adapters' in mapped_columns:
                    try:
                        server.network_adapters = int(float(str(row[mapped_columns['network_adapters']]).replace(',', ''))) if pd.notna(row[mapped_columns['network_adapters']]) else 0
                    except:
                        server.network_adapters = 1  # Default to 1 NIC
                
                if 'recommendation' in mapped_columns:
                    server.recommendation = str(row[mapped_columns['recommendation']]) if pd.notna(row[mapped_columns['recommendation']]) else ""
                
                if 'readiness' in mapped_columns:
                    server.readiness = str(row[mapped_columns['readiness']]) if pd.notna(row[mapped_columns['readiness']]) else ""
                
                if 'estimated_cost' in mapped_columns:
                    try:
                        cost_val = str(row[mapped_columns['estimated_cost']]).replace('$', '').replace(',', '').strip()
                        server.estimated_cost = float(cost_val) if cost_val else 0.0
                    except:
                        server.estimated_cost = 0.0
                
                # Only add server if it has a valid name
                if server.server_name and server.server_name.lower() not in ['nan', 'none', '']:
                    servers.append(server)
                    
            except Exception as e:
                continue  # Skip invalid rows
        
        return servers
    
    @staticmethod
    def _parse_summary_sheet(df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Parse summary data from a sheet."""
        summary = {}
        
        try:
            # Try to extract key-value pairs from the summary sheet
            for _, row in df.iterrows():
                if len(row) >= 2:
                    key = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    value = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                    
                    if key and value and key.lower() not in ['nan', 'none']:
                        summary[key] = value
            
            summary['source_sheet'] = sheet_name
            
        except Exception as e:
            summary['error'] = str(e)
        
        return summary
    
    @staticmethod
    def create_output_excel(excel_output: ExcelOutputType, output_path: str, original_questions_file: str = None):
        """
        Create an Excel file with questions, answers, and analysis matching the example structure.
        
        Args:
            excel_output: ExcelOutputType object with processed data
            output_path: Path where to save the output Excel file
            original_questions_file: Path to original questions file for reference
        """
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            
            # Sheet 1: AI Assisted AIF Completion (matching example structure)
            ws_qa = wb.active
            ws_qa.title = "AI Assisted AIF Completion"
            
            # Headers matching the exact requirements (5 columns only)
            headers = ['Question', 'Answer', 'Confidence', 'Source Reference', 'Status']
            for col, header in enumerate(headers, 1):
                ws_qa.cell(row=1, column=col, value=header)
                ws_qa.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
                # Add header background color
                ws_qa.cell(row=1, column=col).fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Data rows
            for row, qa in enumerate(excel_output.questions_answers, 2):
                ws_qa.cell(row=row, column=1, value=qa.question)
                ws_qa.cell(row=row, column=2, value=qa.answer)
                ws_qa.cell(row=row, column=3, value=qa.confidence)
                ws_qa.cell(row=row, column=4, value=qa.source_reference)
                
                # Determine status more intelligently
                # Consider confidence, source reference, and answer content
                is_actually_answered = (
                    qa.is_answered and 
                    qa.confidence != "Unknown" and 
                    qa.source_reference not in ["N/A", "", "None", None] and
                    qa.answer not in ["Not addressed in transcript", "Error in analysis", "No answer provided", "Not found", ""]
                )
                
                ws_qa.cell(row=row, column=5, value="Answered" if is_actually_answered else "Not Answered")
                
                # Color coding based on confidence
                if qa.confidence == "High":
                    ws_qa.cell(row=row, column=3).fill = openpyxl.styles.PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif qa.confidence == "Medium":
                    ws_qa.cell(row=row, column=3).fill = openpyxl.styles.PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                elif qa.confidence == "Low":
                    ws_qa.cell(row=row, column=3).fill = openpyxl.styles.PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif qa.confidence == "Unknown":
                    ws_qa.cell(row=row, column=3).fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # Color coding for status
                if is_actually_answered:
                    ws_qa.cell(row=row, column=5).fill = openpyxl.styles.PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                else:
                    ws_qa.cell(row=row, column=5).fill = openpyxl.styles.PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Sheet 2: Summary (matching example structure)
            ws_summary = wb.create_sheet(title="Summary")
            
            total_questions = len(excel_output.questions_answers)
            # Calculate actually answered questions using the same intelligent logic
            actually_answered_questions = len([
                qa for qa in excel_output.questions_answers 
                if (qa.is_answered and 
                    qa.confidence != "Unknown" and 
                    qa.source_reference not in ["N/A", "", "None", None] and
                    qa.answer not in ["Not addressed in transcript", "Error in analysis", "No answer provided", "Not found", ""])
            ])
            unanswered_questions = total_questions - actually_answered_questions
            
            # Summary data focusing on core metrics
            summary_data = [
                ["Total Questions", total_questions],
                ["Answered Questions", actually_answered_questions],
                ["Unanswered Questions", unanswered_questions],
                ["Answer Rate", f"{(actually_answered_questions/total_questions*100):.1f}%" if total_questions > 0 else "0%"],
                ["", ""],
                ["Confidence Distribution", ""],
                ["High Confidence", len([qa for qa in excel_output.questions_answers if qa.confidence == "High"])],
                ["Medium Confidence", len([qa for qa in excel_output.questions_answers if qa.confidence == "Medium"])],
                ["Low Confidence", len([qa for qa in excel_output.questions_answers if qa.confidence == "Low"])],
                ["Unknown Confidence", len([qa for qa in excel_output.questions_answers if qa.confidence == "Unknown"])],
                ["", ""],
                ["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ]
            
            for row, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row=row, column=1, value=label)
                ws_summary.cell(row=row, column=2, value=value)
                if label and label != "":
                    ws_summary.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            
            # Sheet 3: Unanswered Questions (matching example structure)
            ws_unanswered = wb.create_sheet(title="Unanswered Questions")
            ws_unanswered.cell(row=1, column=1, value="Unanswered Questions")
            ws_unanswered.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
            ws_unanswered.cell(row=1, column=1).fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            unanswered_count = 2
            for qa in excel_output.questions_answers:
                # Use the same intelligent logic to determine if question is actually unanswered
                is_actually_answered = (
                    qa.is_answered and 
                    qa.confidence != "Unknown" and 
                    qa.source_reference not in ["N/A", "", "None", None] and
                    qa.answer not in ["Not addressed in transcript", "Error in analysis", "No answer provided", "Not found", ""]
                )
                if not is_actually_answered:
                    ws_unanswered.cell(row=unanswered_count, column=1, value=qa.question)
                    unanswered_count += 1
            
            # If no unanswered questions, add a message
            if unanswered_count == 2:
                ws_unanswered.cell(row=2, column=1, value="All questions have been answered!")
                ws_unanswered.cell(row=2, column=1).font = openpyxl.styles.Font(italic=True)
            
            # Auto-adjust column widths
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 80)  # Allow wider columns for questions
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(output_path)
            
        except Exception as e:
            raise Exception(f"Error creating output Excel file: {str(e)}")
    
    @staticmethod
    def validate_excel_file(file_path: str) -> bool:
        """
        Validate if the file is a valid Excel file.
        
        Args:
            file_path: Path to the file
            
        Returns:
            True if valid Excel file, False otherwise
        """
        try:
            pd.read_excel(file_path, nrows=1)
            return True
        except:
            return False
    
    @staticmethod
    def read_dependency_analysis(file_path: str) -> DependencyAnalysis:
        """
        Read and parse Azure Migrate dependency analysis from Excel file.
        
        Args:
            file_path: Path to the dependency analysis Excel file
            
        Returns:
            DependencyAnalysis object with parsed data
        """
        try:
            # Determine engine to use
            file_ext = os.path.splitext(file_path)[1].lower()
            engines = ['openpyxl'] if file_ext == '.xlsx' else ['xlrd'] if file_ext == '.xls' else ['openpyxl', 'xlrd']
            
            for engine in engines:
                try:
                    excel_file = pd.ExcelFile(file_path, engine=engine)
                    break
                except Exception:
                    continue
            else:
                raise Exception(f"Could not read file with any available engine")
            
            connections = []
            network_segments = []
            external_dependencies = []
            internal_dependencies = []
            critical_paths = []
            metadata = {"sheets_processed": [], "file_path": file_path}
            
            # Process each sheet
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    metadata["sheets_processed"].append(sheet_name)
                    
                    # Identify and parse dependency connections
                    if ExcelProcessor._is_dependency_connections_sheet(df, sheet_name):
                        sheet_connections = ExcelProcessor._parse_dependency_connections(df, sheet_name)
                        connections.extend(sheet_connections)
                    
                    # Identify and parse network segments
                    elif ExcelProcessor._is_network_segments_sheet(df, sheet_name):
                        sheet_segments = ExcelProcessor._parse_network_segments(df, sheet_name)
                        network_segments.extend(sheet_segments)
                    
                    # Identify and parse external dependencies
                    elif ExcelProcessor._is_external_dependencies_sheet(df, sheet_name):
                        sheet_external = ExcelProcessor._parse_external_dependencies(df, sheet_name)
                        external_dependencies.extend(sheet_external)
                        
                except Exception as e:
                    metadata[f"error_{sheet_name}"] = str(e)
                    continue
            
            return DependencyAnalysis(
                connections=connections,
                network_segments=network_segments,
                external_dependencies=external_dependencies,
                internal_dependencies=internal_dependencies,
                critical_paths=critical_paths,
                metadata=metadata
            )
            
        except Exception as e:
            raise Exception(f"Error reading dependency analysis file {file_path}: {str(e)}")

    @staticmethod
    def _is_dependency_connections_sheet(df: pd.DataFrame, sheet_name: str) -> bool:
        """Check if the sheet contains dependency connection data."""
        sheet_name_lower = sheet_name.lower()
        dependency_sheet_indicators = [
            'connection', 'dependency', 'dependencies', 'flow', 'communication',
            'network flow', 'server connections', 'app dependencies'
        ]
        
        # Check sheet name
        if any(indicator in sheet_name_lower for indicator in dependency_sheet_indicators):
            return True
        
        # Check column headers
        if df.empty:
            return False
            
        columns_lower = [str(col).lower() for col in df.columns]
        dependency_column_indicators = [
            'source', 'target', 'destination', 'from', 'to', 'protocol', 'port',
            'connection', 'dependency', 'direction', 'flow'
        ]
        
        matches = sum(1 for indicator in dependency_column_indicators 
                     if any(indicator in col for col in columns_lower))
        return matches >= 2
    
    @staticmethod
    def _is_network_segments_sheet(df: pd.DataFrame, sheet_name: str) -> bool:
        """Check if the sheet contains network segment data."""
        sheet_name_lower = sheet_name.lower()
        network_sheet_indicators = ['network', 'segment', 'subnet', 'vlan', 'ip range']
        
        if any(indicator in sheet_name_lower for indicator in network_sheet_indicators):
            return True
        
        if df.empty:
            return False
            
        columns_lower = [str(col).lower() for col in df.columns]
        network_column_indicators = ['subnet', 'vlan', 'segment', 'network', 'ip', 'range']
        
        matches = sum(1 for indicator in network_column_indicators 
                     if any(indicator in col for col in columns_lower))
        return matches >= 1
    
    @staticmethod
    def _is_external_dependencies_sheet(df: pd.DataFrame, sheet_name: str) -> bool:
        """Check if the sheet contains external dependency data."""
        sheet_name_lower = sheet_name.lower()
        external_sheet_indicators = ['external', 'internet', 'cloud', 'third party', 'outside']
        
        return any(indicator in sheet_name_lower for indicator in external_sheet_indicators)

    @staticmethod
    def _parse_dependency_connections(df: pd.DataFrame, sheet_name: str) -> List[DependencyConnection]:
        """Parse dependency connections from DataFrame with focus on network traffic data."""
        connections = []
        
        # Enhanced column mapping to match Azure Migrate dependency analysis format
        column_mapping = {
            'time_slot': ['time slot', 'time', 'timestamp', 'date time', 'period'],
            'source_server': ['source server name', 'source server', 'source machine', 'from server', 'source'],
            'source_ip': ['source ip', 'source ip address', 'source address', 'from ip'],
            'source_application': ['source application', 'source app', 'source process name', 'source service'],
            'source_process': ['source process', 'source proc'],
            'target_server': ['destination server name', 'destination server', 'target server', 'to server', 'destination', 'target'],
            'destination_ip': ['destination ip', 'destination ip address', 'dest ip', 'target ip', 'to ip'],
            'destination_application': ['destination application', 'destination app', 'dest application', 'target application'],
            'destination_process': ['destination process', 'dest process', 'target process'],
            'destination_port': ['destination port', 'dest port', 'target port', 'port', 'service port'],
            'connection_type': ['connection type', 'type', 'service type', 'dependency type', 'service'],
            'protocol': ['protocol', 'transport protocol', 'network protocol'],
            'direction': ['direction', 'flow direction', 'communication direction'],
            'description': ['description', 'details', 'notes', 'comments'],
            'criticality': ['criticality', 'priority', 'importance', 'critical', 'severity']
        }
        
        # Find actual column names using exact matching for Azure Migrate format
        df_columns_lower = [col.lower().strip() for col in df.columns]
        mapped_columns = {}
        
        for standard_name, variations in column_mapping.items():
            for variation in variations:
                if variation in df_columns_lower:
                    mapped_columns[standard_name] = df.columns[df_columns_lower.index(variation)]
                    break
        
        # Process each row
        for _, row in df.iterrows():
            try:
                connection = DependencyConnection()
                
                # Extract all available information using mapped columns
                if 'time_slot' in mapped_columns:
                    connection.time_slot = str(row[mapped_columns['time_slot']]) if pd.notna(row[mapped_columns['time_slot']]) else ""
                
                if 'source_server' in mapped_columns:
                    connection.source_server = str(row[mapped_columns['source_server']]) if pd.notna(row[mapped_columns['source_server']]) else ""
                
                if 'source_ip' in mapped_columns:
                    connection.source_ip = str(row[mapped_columns['source_ip']]) if pd.notna(row[mapped_columns['source_ip']]) else ""
                
                if 'source_application' in mapped_columns:
                    connection.source_application = str(row[mapped_columns['source_application']]) if pd.notna(row[mapped_columns['source_application']]) else ""
                
                if 'source_process' in mapped_columns:
                    connection.source_process = str(row[mapped_columns['source_process']]) if pd.notna(row[mapped_columns['source_process']]) else ""
                
                if 'target_server' in mapped_columns:
                    connection.target_server = str(row[mapped_columns['target_server']]) if pd.notna(row[mapped_columns['target_server']]) else ""
                
                if 'destination_ip' in mapped_columns:
                    connection.destination_ip = str(row[mapped_columns['destination_ip']]) if pd.notna(row[mapped_columns['destination_ip']]) else ""
                
                if 'destination_application' in mapped_columns:
                    connection.destination_application = str(row[mapped_columns['destination_application']]) if pd.notna(row[mapped_columns['destination_application']]) else ""
                
                if 'destination_process' in mapped_columns:
                    connection.destination_process = str(row[mapped_columns['destination_process']]) if pd.notna(row[mapped_columns['destination_process']]) else ""
                
                if 'destination_port' in mapped_columns:
                    connection.destination_port = str(row[mapped_columns['destination_port']]) if pd.notna(row[mapped_columns['destination_port']]) else ""
                    # Also set port for backward compatibility
                    connection.port = connection.destination_port
                
                if 'connection_type' in mapped_columns:
                    connection.connection_type = str(row[mapped_columns['connection_type']]) if pd.notna(row[mapped_columns['connection_type']]) else ""
                
                if 'protocol' in mapped_columns:
                    connection.protocol = str(row[mapped_columns['protocol']]) if pd.notna(row[mapped_columns['protocol']]) else ""
                
                if 'direction' in mapped_columns:
                    connection.direction = str(row[mapped_columns['direction']]) if pd.notna(row[mapped_columns['direction']]) else ""
                
                if 'description' in mapped_columns:
                    connection.description = str(row[mapped_columns['description']]) if pd.notna(row[mapped_columns['description']]) else ""
                
                if 'criticality' in mapped_columns:
                    connection.criticality = str(row[mapped_columns['criticality']]) if pd.notna(row[mapped_columns['criticality']]) else ""
                
                # Add connection if we have essential network traffic data (source and destination info)
                if (connection.source_ip and connection.destination_ip) or (connection.source_server and connection.target_server):
                    connections.append(connection)
                    
            except Exception as e:
                print(f"Warning: Error parsing connection row: {e}")
                continue
        
        return connections

    @staticmethod
    def _parse_network_segments(df: pd.DataFrame, sheet_name: str) -> List[NetworkSegment]:
        """Parse network segment data from a sheet."""
        segments = []
        
        # Column mapping for network segments
        column_mapping = {
            'segment_name': ['segment name', 'network name', 'subnet name', 'name'],
            'subnet': ['subnet', 'ip range', 'network range', 'cidr', 'network'],
            'vlan_id': ['vlan', 'vlan id', 'vlan number'],
            'purpose': ['purpose', 'description', 'type', 'function', 'role'],
            'servers': ['servers', 'hosts', 'machines', 'computers', 'devices']
        }
        
        # Find actual column names
        df_columns_lower = [col.lower().strip() for col in df.columns]
        mapped_columns = {}
        
        for standard_name, variations in column_mapping.items():
            for variation in variations:
                if variation in df_columns_lower:
                    mapped_columns[standard_name] = df.columns[df_columns_lower.index(variation)]
                    break
        
        # Process each row
        for _, row in df.iterrows():
            try:
                segment = NetworkSegment()
                
                if 'segment_name' in mapped_columns:
                    segment.segment_name = str(row[mapped_columns['segment_name']]) if pd.notna(row[mapped_columns['segment_name']]) else ""
                
                if 'subnet' in mapped_columns:
                    segment.subnet = str(row[mapped_columns['subnet']]) if pd.notna(row[mapped_columns['subnet']]) else ""
                
                if 'vlan_id' in mapped_columns:
                    segment.vlan_id = str(row[mapped_columns['vlan_id']]) if pd.notna(row[mapped_columns['vlan_id']]) else ""
                
                if 'purpose' in mapped_columns:
                    segment.purpose = str(row[mapped_columns['purpose']]) if pd.notna(row[mapped_columns['purpose']]) else ""
                
                if 'servers' in mapped_columns and pd.notna(row[mapped_columns['servers']]):
                    # Split server list if it's comma-separated
                    servers_str = str(row[mapped_columns['servers']])
                    segment.servers = [s.strip() for s in servers_str.split(',') if s.strip()]
                
                segments.append(segment)
                    
            except Exception as e:
                print(f"Warning: Error parsing network segment row: {e}")
                continue
        
        return segments

    @staticmethod
    def _parse_external_dependencies(df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Parse external dependency data from a sheet."""
        external_deps = []
        
        # Look for columns that might contain external dependencies
        for column in df.columns:
            if any(keyword in column.lower() for keyword in ['external', 'dependency', 'service', 'url', 'endpoint']):
                for _, row in df.iterrows():
                    if pd.notna(row[column]) and str(row[column]).strip():
                        external_deps.append(str(row[column]).strip())
        
        return list(set(external_deps))  # Remove duplicates
